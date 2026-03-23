"""
User Authentication Views for CTRG Grant System

This module provides REST API endpoints for user authentication, registration,
and user management. It implements token-based authentication using Django
REST Framework's built-in token authentication.

Views:
    - LoginView: User login with email/password
    - LogoutView: User logout (token destruction)
    - CurrentUserView: Get authenticated user's profile
    - UserRegistrationView: Create new users (admin only)
    - ChangePasswordView: Change user password
    - UserListView: List all users (admin only)

Authentication Method: Token-based (DRF AuthToken)
"""

from rest_framework import status, generics, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework.throttling import AnonRateThrottle
from django.contrib.auth import get_user_model
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from django.utils.crypto import get_random_string
from proposals.models import AuditLog
from proposals.services import get_client_ip


class LoginRateThrottle(AnonRateThrottle):
    rate = '5/minute'


class RegistrationRateThrottle(AnonRateThrottle):
    rate = '10/hour'

from .serializers import (
    UserSerializer,
    LoginSerializer,
    UserCreateSerializer,
    ChangePasswordSerializer,
    UserListSerializer,
    ReviewerRegistrationSerializer
)

# Get the custom User model
User = get_user_model()


def _normalize_header(value):
    return str(value).strip().lower().replace(' ', '_') if value is not None else ''


def _get_primary_role(user):
    """Return user's primary role from group membership."""
    if user.groups.exists():
        return user.groups.first().name
    if user.is_staff:
        return 'SRC_Chair'
    return None


def _get_role_redirect_path(user, role=None):
    """Return default frontend redirect path for the user's role."""
    resolved_role = role if role is not None else _get_primary_role(user)
    normalized = (resolved_role or '').strip().lower()
    if user.is_staff or normalized in {'src_chair', 'src chair', 'admin'}:
        return '/admin/dashboard'
    if normalized == 'reviewer':
        return '/reviewer/dashboard'
    if normalized == 'pi':
        return '/pi/dashboard'
    return '/unauthorized'


def _generate_unique_username(email):
    base = email.split('@')[0].strip().lower().replace(' ', '.') or 'reviewer'
    candidate = base
    suffix = 1
    while User.objects.filter(username=candidate).exists():
        candidate = f"{base}{suffix}"
        suffix += 1
    return candidate


def _generate_temp_password():
    # Meets minimum length and avoids common-password/numeric-only failures.
    return f"Rvwr!{get_random_string(10)}"


def _audit_user_event(request, action_type, actor=None, target_user=None, details=None):
    """Write auth and user-management events into the shared audit trail."""
    AuditLog.objects.create(
        user=actor,
        proposal=None,
        action_type=action_type,
        details={
            'actor_email': actor.email if actor else None,
            'target_user_id': target_user.id if target_user else None,
            'target_user_email': target_user.email if target_user else None,
            **(details or {}),
        },
        ip_address=get_client_ip(request),
    )


class LoginView(ObtainAuthToken):
    """
    User login endpoint that returns authentication token and user details.

    POST /api/auth/login/

    Accepts email and password, validates credentials, and returns:
    - Authentication token (for subsequent API requests)
    - User role (PI, Reviewer, or SRC_Chair)
    - Complete user profile

    Request Body:
        {
            "email": "user@nsu.edu",
            "password": "<your_password>"
        }

    Success Response (200 OK):
        {
            "access": "a1b2c3d4e5f6...",  # Auth token
            "role": "SRC_Chair",
            "user": {
                "id": 1,
                "username": "john.doe",
                "email": "user@nsu.edu",
                "first_name": "John",
                "last_name": "Doe",
                "is_active": true
            }
        }

    Error Responses:
        - 400 Bad Request: Invalid email format or missing fields
        - 401 Unauthorized: Invalid credentials or inactive account

    Authentication: Not required (public endpoint)
    """

    serializer_class = LoginSerializer
    permission_classes = [permissions.AllowAny]
    throttle_classes = [LoginRateThrottle]

    def post(self, request, *args, **kwargs):
        """
        Authenticate user and return token with user details.

        Args:
            request: HTTP request with email and password

        Returns:
            Response: Authentication token, role, and user profile
        """
        # Validate credentials using LoginSerializer
        serializer = self.serializer_class(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)

        # Get authenticated user from serializer validation
        user = serializer.validated_data['user']

        # Get or create authentication token for this user
        token, created = Token.objects.get_or_create(user=user)

        # Determine user's role from group membership
        role = _get_primary_role(user)

        # Serialize user data for response
        user_serializer = UserSerializer(user)

        # Return token, role, and user details
        _audit_user_event(
            request,
            action_type='USER_LOGIN',
            actor=user,
            target_user=user,
            details={'role': role, 'token_created': created},
        )
        return Response({
            'access': token.key,  # Named 'access' for frontend compatibility
            'role': role,
            'redirect_to': _get_role_redirect_path(user, role=role),
            'user': user_serializer.data
        }, status=status.HTTP_200_OK)


class LogoutView(APIView):
    """
    User logout endpoint that destroys the authentication token.

    POST /api/auth/logout/

    Deletes the user's authentication token, effectively logging them out.
    After logout, the token cannot be used for API authentication.

    Request Body: Empty (authentication via token in header)

    Success Response (200 OK):
        {
            "message": "Successfully logged out."
        }

    Error Responses:
        - 401 Unauthorized: No valid token provided

    Authentication: Required (Token)
    """

    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """
        Delete user's authentication token.

        Args:
            request: HTTP request with authentication token

        Returns:
            Response: Success message
        """
        try:
            # Delete the user's token
            request.user.auth_token.delete()
            _audit_user_event(
                request,
                action_type='USER_LOGOUT',
                actor=request.user,
                target_user=request.user,
            )
            return Response(
                {'message': 'Successfully logged out.'},
                status=status.HTTP_200_OK
            )
        except Exception as e:
            # Handle edge case where token doesn't exist
            return Response(
                {'error': 'Logout failed.'},
                status=status.HTTP_400_BAD_REQUEST
            )


class CurrentUserView(APIView):
    """
    Get current authenticated user's profile information.

    GET /api/auth/user/

    Returns complete profile information for the currently authenticated user.
    Used by frontend to fetch user details after login or on page reload.

    Success Response (200 OK):
        {
            "id": 1,
            "username": "john.doe",
            "email": "john.doe@nsu.edu",
            "first_name": "John",
            "last_name": "Doe",
            "role": "SRC_Chair",
            "is_active": true
        }

    Error Responses:
        - 401 Unauthorized: No valid token provided

    Authentication: Required (Token)
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """
        Return current user's profile data.

        Args:
            request: HTTP request with authentication token

        Returns:
            Response: User profile data
        """
        # Serialize authenticated user's data
        serializer = UserSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TokenValidationView(APIView):
    """
    Validate token and return user role metadata for role-based redirection.

    GET /api/auth/validate-token/
    """

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user_serializer = UserSerializer(request.user)
        role = _get_primary_role(request.user)
        return Response({
            'valid': True,
            'role': role,
            'redirect_to': _get_role_redirect_path(request.user, role=role),
            'user': user_serializer.data
        }, status=status.HTTP_200_OK)


class UserRegistrationView(generics.CreateAPIView):
    """
    Create new user account (Admin only).

    POST /api/auth/register/

    Allows SRC Chair (admin) to create new user accounts for PIs, Reviewers,
    or other SRC Chairs. Handles password hashing and role assignment.

    Request Body:
        {
            "username": "jane.smith",
            "email": "jane.smith@nsu.edu",
            "password": "SecurePass123!",
            "first_name": "Jane",
            "last_name": "Smith",
            "role": "Reviewer"  // Must be: PI, Reviewer, or SRC_Chair
        }

    Success Response (201 Created):
        {
            "id": 5,
            "username": "jane.smith",
            "email": "jane.smith@nsu.edu",
            "first_name": "Jane",
            "last_name": "Smith",
            "role": "Reviewer",
            "is_active": true
        }

    Error Responses:
        - 400 Bad Request: Invalid data, duplicate email/username, weak password
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user

    Authentication: Required (Token)
    Permissions: Admin users only (is_staff=True)
    """

    serializer_class = UserCreateSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def perform_create(self, serializer):
        """
        Create user and log the creation action.

        Args:
            serializer: Validated UserCreateSerializer
        """
        # Save the new user (serializer handles password hashing and role assignment)
        user = serializer.save()

        _audit_user_event(
            self.request,
            action_type='USER_CREATED',
            actor=self.request.user,
            target_user=user,
            details={'role': _get_primary_role(user)},
        )


class ImportReviewersFromExcelView(APIView):
    """
    Import reviewer accounts from an Excel file (.xlsx).

    Expected header columns:
    - email (required)
    - first_name (required)
    - last_name (required)
    - username (optional, auto-generated if missing)
    - password (optional, temporary password auto-generated if missing)
    """

    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'No file uploaded. Provide file in form-data key "file".'}, status=status.HTTP_400_BAD_REQUEST)

        filename = (upload.name or '').lower()
        if not filename.endswith('.xlsx'):
            return Response({'error': 'Unsupported file type. Please upload an .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(filename=upload, data_only=True)
            sheet = workbook.active
        except Exception:
            return Response({'error': 'Unable to read Excel file. Ensure the file is a valid .xlsx workbook.'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'The Excel file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        header = [_normalize_header(cell) for cell in rows[0]]
        aliases = {
            'first_name': {'first_name', 'firstname', 'first'},
            'last_name': {'last_name', 'lastname', 'last'},
            'email': {'email', 'email_address', 'mail'},
            'username': {'username', 'user_name', 'user'},
            'password': {'password', 'pass', 'temp_password'},
        }

        index_map = {}
        for field, names in aliases.items():
            for idx, col in enumerate(header):
                if col in names:
                    index_map[field] = idx
                    break

        required = ['first_name', 'last_name', 'email']
        missing = [field for field in required if field not in index_map]
        if missing:
            return Response(
                {'error': f"Missing required columns: {', '.join(missing)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        created = []
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            def _cell(field):
                idx = index_map.get(field)
                if idx is None or idx >= len(row):
                    return ''
                return str(row[idx]).strip() if row[idx] is not None else ''

            first_name = _cell('first_name')
            last_name = _cell('last_name')
            email = _cell('email')
            username = _cell('username') or _generate_unique_username(email)
            password = _cell('password') or _generate_temp_password()

            payload = {
                'username': username,
                'email': email,
                'password': password,
                'first_name': first_name,
                'last_name': last_name,
                'role': 'Reviewer',
            }

            serializer = UserCreateSerializer(data=payload)
            if serializer.is_valid():
                user = serializer.save()
                created_row = {
                    'row': row_idx,
                    'id': user.id,
                    'email': user.email,
                    'username': user.username,
                }
                if 'password' not in index_map or not _cell('password'):
                    created_row['has_temporary_password'] = True
                    # Send temp password via email instead of returning in response
                    try:
                        from django.core.mail import send_mail
                        from django.conf import settings
                        send_mail(
                            subject='CTRG Grant System - Your Account Credentials',
                            message=f"Dear {first_name},\n\nYour reviewer account has been created.\n\nUsername: {username}\nEmail: {email}\nTemporary Password: {password}\n\nPlease change your password after logging in.\n\nBest regards,\nCTRG Grant Review System",
                            from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@nsu.edu'),
                            recipient_list=[email],
                            fail_silently=True,
                        )
                    except Exception:
                        pass
                created.append(created_row)
                _audit_user_event(
                    request,
                    action_type='REVIEWER_IMPORTED',
                    actor=request.user,
                    target_user=user,
                    details={'row': row_idx, 'temporary_password_emailed': created_row.get('has_temporary_password', False)},
                )
            else:
                errors.append({
                    'row': row_idx,
                    'email': email,
                    'errors': serializer.errors,
                })

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        }, status=status.HTTP_200_OK)


class ChangePasswordView(APIView):
    """
    Change user password.

    POST /api/auth/change-password/

    Allows authenticated users to change their password by providing their
    current password for verification and a new password.

    Request Body:
        {
            "old_password": "OldPass123!",
            "new_password": "NewSecurePass456!"
        }

    Success Response (200 OK):
        {
            "message": "Password successfully changed."
        }

    Error Responses:
        - 400 Bad Request: Invalid old password, weak new password, or same password
        - 401 Unauthorized: Not authenticated

    Authentication: Required (Token)

    Note: After password change, the auth token remains valid. Consider
    invalidating the token in production to force re-login.
    """

    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        """
        Validate and update user's password.

        Args:
            request: HTTP request with old and new passwords

        Returns:
            Response: Success message or validation errors
        """
        # Validate password change request
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)

        # Save new password (serializer handles hashing)
        serializer.save()
        _audit_user_event(
            request,
            action_type='PASSWORD_CHANGED',
            actor=request.user,
            target_user=request.user,
        )

        return Response(
            {'message': 'Password successfully changed.'},
            status=status.HTTP_200_OK
        )


class UserListView(generics.ListAPIView):
    """
    List all users in the system (Admin only).

    GET /api/auth/users/

    Returns a list of all user accounts with basic information.
    Used by admin dashboard to view and manage users.

    Query Parameters:
        - role: Filter by role (optional) - e.g., ?role=Reviewer
        - is_active: Filter by active status (optional) - e.g., ?is_active=true

    Success Response (200 OK):
        [
            {
                "id": 1,
                "username": "john.doe",
                "email": "john.doe@nsu.edu",
                "full_name": "John Doe",
                "role": "SRC_Chair",
                "is_active": true
            },
            {
                "id": 2,
                "username": "jane.smith",
                "email": "jane.smith@nsu.edu",
                "full_name": "Jane Smith",
                "role": "Reviewer",
                "is_active": true
            }
        ]

    Error Responses:
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user

    Authentication: Required (Token)
    Permissions: Admin users only (is_staff=True)
    """

    serializer_class = UserListSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    queryset = User.objects.all().order_by('-date_joined')

    def get_queryset(self):
        """
        Get filtered user queryset based on query parameters.

        Supports filtering by:
        - role: User's group/role name
        - is_active: Active status

        Returns:
            QuerySet: Filtered user queryset
        """
        queryset = super().get_queryset()

        # Filter by role if provided
        role = self.request.query_params.get('role', None)
        if role:
            queryset = queryset.filter(groups__name=role)

        # Filter by active status if provided
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            is_active_bool = is_active.lower() == 'true'
            queryset = queryset.filter(is_active=is_active_bool)

        return queryset


# Additional view for user detail/update/delete (optional enhancement)
class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update, or delete a specific user (Admin only).

    GET    /api/auth/users/<id>/  - Get user details
    PUT    /api/auth/users/<id>/  - Update user
    PATCH  /api/auth/users/<id>/  - Partial update user
    DELETE /api/auth/users/<id>/  - Delete user (soft delete recommended)

    Success Response (200 OK for GET/PUT/PATCH, 204 No Content for DELETE)

    Error Responses:
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user
        - 404 Not Found: User does not exist

    Authentication: Required (Token)
    Permissions: Admin users only (is_staff=True)
    """

    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]
    queryset = User.objects.all()
    lookup_field = 'pk'

    def perform_update(self, serializer):
        user = serializer.save()
        _audit_user_event(
            self.request,
            action_type='USER_UPDATED',
            actor=self.request.user,
            target_user=user,
            details={'role': _get_primary_role(user), 'is_active': user.is_active},
        )

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        from reviews.models import ReviewAssignment
        active_assignments = ReviewAssignment.objects.filter(reviewer=user).exists()
        if active_assignments:
            # Soft delete: deactivate instead of deleting
            user.is_active = False
            user.save(update_fields=['is_active'])
            _audit_user_event(
                request,
                action_type='USER_DEACTIVATED',
                actor=request.user,
                target_user=user,
                details={'reason': 'active_review_assignments'},
            )
            return Response(
                {'message': 'User has review assignments. Account has been deactivated instead of deleted.'},
                status=status.HTTP_200_OK
            )
        _audit_user_event(
            request,
            action_type='USER_DELETED',
            actor=request.user,
            target_user=user,
        )
        return super().destroy(request, *args, **kwargs)



class PendingReviewersView(generics.ListAPIView):
    """
    List all pending (inactive) reviewer registrations (Admin only).

    GET /api/auth/pending-reviewers/

    Returns a list of all pending reviewer accounts awaiting SRC Chair approval.
    Used by admin dashboard to review and approve new reviewer registrations.

    Success Response (200 OK):
        [
            {
                "id": 1,
                "username": "john.reviewer",
                "email": "john.reviewer@nsu.edu",
                "full_name": "John Reviewer",
                "role": "Reviewer",
                "is_active": false,
                "date_joined": "2024-02-09T10:00:00Z"
            }
        ]

    Error Responses:
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user

    Authentication: Required (Token)
    Permissions: Admin users only (is_staff=True)
    """

    serializer_class = UserListSerializer
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def get_queryset(self):
        """
        Get all pending (inactive) reviewers.

        Returns:
            QuerySet: Inactive users in the Reviewer group
        """
        return User.objects.filter(
            groups__name='Reviewer',
            is_active=False
        ).order_by('-date_joined')


class ReviewerCVDownloadView(APIView):
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def get(self, request, pk):
        user = get_object_or_404(User, pk=pk, groups__name='Reviewer')
        profile = getattr(user, 'reviewer_profile', None)
        if not profile or not profile.cv:
            return Response(
                {'error': 'CV not found for this reviewer.'},
                status=status.HTTP_404_NOT_FOUND
            )

        return FileResponse(
            profile.cv.open('rb'),
            as_attachment=False,
            filename=profile.cv.name.rsplit('/', 1)[-1]
        )


class ApproveReviewerView(APIView):
    """
    ============================================================================
    APPROVE PENDING REVIEWER REGISTRATION
    ============================================================================

    PURPOSE:
    Activates a pending reviewer account, allowing them to login and review
    proposals.

    ENDPOINT: POST /api/auth/approve-reviewer/<id>/

    WHAT IT DOES:
    1. Validates user exists and is a pending reviewer
    2. Sets User.is_active = True (enables login)
    3. Sets ReviewerProfile.is_active_reviewer = True (enables assignments)
    4. Returns success message with user data

    VALIDATION:
    - User must exist
    - User must be in "Reviewer" group
    - User must be inactive (is_active=False)
    - Cannot approve already active reviewers

    PERMISSIONS:
    - Requires authentication (token)
    - Requires admin status (is_staff=True or SRC_Chair group)

    SUCCESS RESPONSE (200 OK):
        {
            "message": "Reviewer approved successfully.",
            "user": {
                "id": 1,
                "username": "john.reviewer",
                "email": "john.reviewer@nsu.edu",
                "first_name": "John",
                "last_name": "Reviewer",
                "is_active": true,
                "role": "Reviewer"
            }
        }

    ERROR RESPONSES:
        - 400 Bad Request: User is already active or not a reviewer
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user
        - 404 Not Found: User does not exist

    EXAMPLE:
        POST /api/auth/approve-reviewer/5/
        Authorization: Token abc123...

        Response: {"message": "Reviewer approved successfully.", "user": {...}}
    """

    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def post(self, request, pk):
        """
        Approve pending reviewer and activate their account.

        WORKFLOW:
        1. Fetch user by ID
        2. Validate user is in Reviewer group
        3. Check if already active (prevent duplicate approvals)
        4. Set User.is_active = True
        5. Set ReviewerProfile.is_active_reviewer = True
        6. Return success response

        Args:
            request: HTTP request with authentication token
            pk: User ID (primary key)

        Returns:
            Response: Success message and serialized user data (200 OK)
                     OR error message (400/404)
        """
        # ====================================================================
        # STEP 1: Fetch user and validate existence
        # ====================================================================
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # ====================================================================
        # STEP 2: Validate user is in Reviewer group
        # ====================================================================
        # Safety check: Only approve users who are actually reviewers
        if not user.groups.filter(name='Reviewer').exists():
            return Response(
                {'error': 'User is not a reviewer.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ====================================================================
        # STEP 3: Check if already active
        # ====================================================================
        # Prevent duplicate approvals (idempotency check)
        if user.is_active:
            return Response(
                {'error': 'Reviewer is already approved.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ====================================================================
        # STEP 4: Activate user account
        # ====================================================================
        # Setting is_active=True enables login functionality
        user.is_active = True
        user.save()

        # ====================================================================
        # STEP 5: Activate reviewer profile
        # ====================================================================
        # ReviewerProfile must also be activated to receive review assignments
        try:
            from reviews.models import ReviewerProfile
            reviewer_profile = ReviewerProfile.objects.get(user=user)
            reviewer_profile.is_active_reviewer = True
            reviewer_profile.save()
        except ReviewerProfile.DoesNotExist:
            # Profile doesn't exist - this shouldn't happen in normal flow
            # (created automatically during registration)
            # Handle gracefully - approval still succeeds
            pass

        # ====================================================================
        # STEP 6: Return success response
        # ====================================================================
        serializer = UserSerializer(user)
        _audit_user_event(
            request,
            action_type='REVIEWER_APPROVED',
            actor=request.user,
            target_user=user,
        )
        return Response({
            'message': 'Reviewer approved successfully.',
            'user': serializer.data
        }, status=status.HTTP_200_OK)


class RejectReviewerView(APIView):
    """
    ============================================================================
    REJECT PENDING REVIEWER REGISTRATION
    ============================================================================

    PURPOSE:
    Permanently deletes a pending reviewer account, rejecting their registration.

    ENDPOINT: DELETE /api/auth/reject-reviewer/<id>/

    WHAT IT DOES:
    1. Validates user exists and is a pending reviewer
    2. PERMANENTLY DELETES the user account
    3. CASCADE DELETES associated ReviewerProfile
    4. Returns success message

    ⚠️ WARNING: This is a DESTRUCTIVE operation!
    - User account is permanently deleted
    - Cannot be undone
    - User must re-register if rejected by mistake

    VALIDATION & SAFETY:
    - User must exist
    - User must be in "Reviewer" group
    - User must be INACTIVE (is_active=False)
    - CANNOT reject active reviewers (safety check)

    PERMISSIONS:
    - Requires authentication (token)
    - Requires admin status (is_staff=True or SRC_Chair group)

    SUCCESS RESPONSE (200 OK):
        {
            "message": "Reviewer registration rejected."
        }

    ERROR RESPONSES:
        - 400 Bad Request: User is active or not a reviewer
        - 401 Unauthorized: Not authenticated
        - 403 Forbidden: Not an admin user
        - 404 Not Found: User does not exist

    EXAMPLE:
        DELETE /api/auth/reject-reviewer/5/
        Authorization: Token abc123...

        Response: {"message": "Reviewer registration rejected."}
    """

    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def delete(self, request, pk):
        """
        Reject pending reviewer registration by deleting the account.

        WORKFLOW:
        1. Fetch user by ID
        2. Validate user is in Reviewer group
        3. Check user is inactive (prevent deleting active reviewers)
        4. Delete user account (CASCADE deletes ReviewerProfile)
        5. Return success response

        Args:
            request: HTTP request with authentication token
            pk: User ID (primary key)

        Returns:
            Response: Success message (200 OK) OR error message (400/404)
        """
        # ====================================================================
        # STEP 1: Fetch user and validate existence
        # ====================================================================
        try:
            user = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # ====================================================================
        # STEP 2: Validate user is in Reviewer group
        # ====================================================================
        # Safety check: Only reject reviewer accounts
        if not user.groups.filter(name='Reviewer').exists():
            return Response(
                {'error': 'User is not a reviewer.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ====================================================================
        # STEP 3: CRITICAL SAFETY CHECK - Prevent deletion of active users
        # ====================================================================
        # This prevents accidentally deleting reviewers who are already
        # approved and working in the system
        if user.is_active:
            return Response(
                {'error': 'Cannot reject an active reviewer. Use the deactivate endpoint instead.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ====================================================================
        # STEP 4: DELETE user account (DESTRUCTIVE - Cannot be undone!)
        # ====================================================================
        _audit_user_event(
            request,
            action_type='REVIEWER_REJECTED',
            actor=request.user,
            target_user=user,
        )

        # Django's cascade deletion will also delete:
        # - ReviewerProfile (ForeignKey to User)
        # - Any other related objects
        user.delete()

        # ====================================================================
        # STEP 5: Return success response
        # ====================================================================
        return Response(
            {'message': 'Reviewer registration rejected.'},
            status=status.HTTP_200_OK
        )


class InviteReviewerView(APIView):
    """
    SRC Chair invites a reviewer by email. Sends an invitation link with a unique token.

    POST /api/auth/invite-reviewer/
    """
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def post(self, request):
        from .serializers import ReviewerInvitationSerializer
        from .models import ReviewerInvitation
        from datetime import timedelta
        from django.utils import timezone
        from django.core.mail import send_mail
        from django.conf import settings

        serializer = ReviewerInvitationSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        email = serializer.validated_data['email']
        expires_in_days = serializer.validated_data.get('expires_in_days', 7)

        # Cancel any previous unused invitations for this email (not "used", just cancelled)
        ReviewerInvitation.objects.filter(email=email, is_used=False).delete()

        invitation = ReviewerInvitation.objects.create(
            email=email,
            invited_by=request.user,
            expires_at=timezone.now() + timedelta(days=expires_in_days),
        )

        # Build the registration URL using FRONTEND_URL setting
        frontend_origin = getattr(settings, 'FRONTEND_URL', '').rstrip('/')
        if not frontend_origin:
            frontend_origin = request.META.get('HTTP_ORIGIN', 'http://localhost:5173')

        registration_url = f"{frontend_origin}/register-reviewer?token={invitation.token}"

        # Send invitation email
        subject = "CTRG - Invitation to Register as Reviewer"
        message = (
            f"Dear Colleague,\n\n"
            f"You have been invited by the SRC Chair to register as a reviewer "
            f"for the CTRG Grant Review System at NSU.\n\n"
            f"Please use the following link to complete your registration:\n"
            f"{registration_url}\n\n"
            f"This invitation will expire in {expires_in_days} day(s).\n\n"
            f"If you did not expect this invitation, please disregard this email.\n\n"
            f"Best regards,\n"
            f"CTRG Grant Review System"
        )

        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            email_sent = True
        except Exception:
            email_sent = False

        _audit_user_event(
            request,
            action_type='REVIEWER_INVITED',
            actor=request.user,
            details={'invited_email': email, 'invitation_id': invitation.id, 'email_sent': email_sent},
        )

        # Always return the registration_url to the SRC Chair.
        # This is an authenticated admin-only endpoint. The chair needs the URL
        # to share manually if email delivery fails (e.g. no mail server configured).
        # The token is single-use and time-limited, so returning it to the
        # authenticated admin is intentional and safe.
        response_data = {
            'message': f'Invitation sent to {email}.',
            'email_sent': email_sent,
            'expires_at': invitation.expires_at.isoformat(),
            'registration_url': registration_url,
        }

        return Response(response_data, status=status.HTTP_201_CREATED)


class ValidateInvitationTokenView(APIView):
    """
    Validate an invitation token and return the invited email.

    GET /api/auth/validate-invitation/<token>/
    """
    permission_classes = [permissions.AllowAny]
    throttle_classes = [RegistrationRateThrottle]

    def get(self, request, token):
        from .models import ReviewerInvitation

        try:
            invitation = ReviewerInvitation.objects.get(token=token)
        except (ReviewerInvitation.DoesNotExist, ValueError):
            return Response(
                {'valid': False, 'error': 'Invalid invitation token.'},
                status=status.HTTP_404_NOT_FOUND
            )

        if invitation.is_used:
            return Response(
                {'valid': False, 'error': 'This invitation has already been used.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if invitation.is_expired:
            return Response(
                {'valid': False, 'error': 'This invitation has expired.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'valid': True,
            'email': invitation.email,
            'expires_at': invitation.expires_at.isoformat(),
        })


class InvitedReviewerRegistrationView(generics.CreateAPIView):
    """
    Register a reviewer using a valid invitation token.

    POST /api/auth/register-reviewer/
    Replaces the old public self-registration. Requires invitation token.
    """
    permission_classes = [permissions.AllowAny]
    throttle_classes = [RegistrationRateThrottle]

    def get_serializer_class(self):
        from .serializers import InvitedReviewerRegistrationSerializer
        return InvitedReviewerRegistrationSerializer

    def perform_create(self, serializer):
        user = serializer.save()
        _audit_user_event(
            self.request,
            action_type='REVIEWER_REGISTERED_VIA_INVITATION',
            actor=user,
            target_user=user,
            details={'role': _get_primary_role(user)},
        )


class ListInvitationsView(generics.ListAPIView):
    """
    List all reviewer invitations (Admin only).

    GET /api/auth/invitations/
    """
    permission_classes = [permissions.IsAuthenticated, permissions.IsAdminUser]

    def get(self, request):
        from .models import ReviewerInvitation
        invitations = ReviewerInvitation.objects.all().order_by('-created_at')
        data = []
        for inv in invitations:
            data.append({
                'id': inv.id,
                'email': inv.email,
                # Raw token is omitted to prevent anyone with read access to
                # the admin API (or its logs) from obtaining a usable token.
                'invited_by': inv.invited_by.get_full_name() if inv.invited_by else None,
                'created_at': inv.created_at.isoformat(),
                'expires_at': inv.expires_at.isoformat(),
                'is_used': inv.is_used,
                'is_expired': inv.is_expired,
                'is_valid': inv.is_valid,
            })
        return Response(data)


class TestEmailView(APIView):
    """
    Send a test email to verify SMTP configuration.
    Admin only. POST with optional {"recipient": "email@example.com"}.
    """
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        """Return current email backend configuration (no secrets)."""
        from django.conf import settings
        backend = getattr(settings, 'EMAIL_BACKEND', '')
        is_console = 'console' in backend.lower()
        return Response({
            'backend': backend,
            'is_smtp': not is_console,
            'host': getattr(settings, 'EMAIL_HOST', ''),
            'port': getattr(settings, 'EMAIL_PORT', 587),
            'use_tls': getattr(settings, 'EMAIL_USE_TLS', True),
            'from_email': getattr(settings, 'DEFAULT_FROM_EMAIL', ''),
            'user_configured': bool(getattr(settings, 'EMAIL_HOST_USER', '')),
        })

    def post(self, request):
        """Send a test email."""
        from django.core.mail import send_mail
        from django.conf import settings

        recipient = request.data.get('recipient') or request.user.email
        if not recipient:
            return Response({'error': 'No recipient email provided.'}, status=400)

        backend = getattr(settings, 'EMAIL_BACKEND', '')
        is_console = 'console' in backend.lower()

        try:
            sent = send_mail(
                subject='CTRG System — Test Email',
                message=(
                    'This is a test email from the CTRG Grant Review Management System.\n\n'
                    'If you received this, your email configuration is working correctly.\n\n'
                    'Best regards,\nCTRG Grant Review System\nNorth South University'
                ),
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[recipient],
                fail_silently=False,
            )
            return Response({
                'success': True,
                'sent_to': recipient,
                'backend': backend,
                'note': 'Email printed to server console (no SMTP configured).' if is_console else 'Email sent via SMTP.',
            })
        except Exception as exc:
            return Response({'success': False, 'error': str(exc)}, status=500)
